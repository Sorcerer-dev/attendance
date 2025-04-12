# excel_utils.py
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def save_attendance_to_excel(file_name, selected_date, button_states):
    data = {"Roll No": list(range(1, 61)), selected_date: button_states}
    df_new = pd.DataFrame(data)

    if os.path.exists(file_name):
        df = pd.read_excel(file_name)

        if "Roll No" not in df.columns:
            df.insert(0, "Roll No", list(range(1, 61)))

        if selected_date in df.columns:
            df[selected_date] = df_new[selected_date]
        else:
            df = df.merge(df_new, on="Roll No", how="outer")
    else:
        df = df_new

    date_columns = [col for col in df.columns if col != "Roll No"]
    sorted_dates = sorted(date_columns, key=lambda x: datetime.strptime(x, "%d-%m-%Y"))
    df = df[["Roll No"] + sorted_dates]

    df.to_excel(file_name, index=False)
    format_excel(file_name)

def format_excel(file_name):
    wb = load_workbook(file_name)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, min_col=2):
        for cell in row:
            val = str(cell.value).strip()
            if val == "P":
                cell.style = "Good"
            elif val == "A":
                cell.style = "Bad"
            elif val == "H":
                cell.style = "Neutral"
            elif val == "OD":
                cell.style = "Check Cell"
            cell.alignment = Alignment(horizontal='center', vertical='center')

    for cell in ws["A"]:
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
        for cell in col:
            ws.column_dimensions[cell.column_letter].width = 15
            break

    wb.save(file_name)
